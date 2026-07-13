"""
Streamlit app: merge Taiwan customs trade CSV exports into one formatted Excel file.

Workflow implemented:
  1. Accepts one or more CSV files (drag/drop or browse).
  2. Drops 中文貨名 and 英文貨名 columns.
  3. Reorders columns to: 進出口別 | 日期 | 貨品號列 | 國家 | 重量(公斤) | 新臺幣(千元)
  4. Replaces 進口 -> IMPORT and 出口 -> EXPORT.
  5. Converts ROC-era dates ("113年05月" or "113年05月01日") to real Excel dates
     formatted as YYYY/MM/DD.
  6. Forces 貨品號列 to a true numeric type so it works in Excel formulas.
  7. Merges all uploaded files, sorted by year-month ascending only (day/time
     is ignored); within the same month, IMPORT rows are placed above EXPORT
     rows and all other original ordering is preserved.
  8. Detects which month(s) are present and lets you enter an exchange rate
     for each; that rate is applied to every row belonging to that month in
     a new rightmost 匯率 column.
  9. Outputs a polished .xlsx (bold header, frozen header row, autofit columns,
     proper number formats) and a one-click "copy to clipboard" option
     (tab-separated, no header row) for pasting straight into an existing sheet.

Run with:  streamlit run trade_data_app.py
"""

import io
import json
import re
from datetime import datetime

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Constants
# --------------------------------------------------------------------------

FINAL_COLUMNS = ["進出口別", "日期", "貨品號列", "國家", "重量(公斤)", "新臺幣(千元)"]
EXCHANGE_COL = "匯率"
OUTPUT_COLUMNS = FINAL_COLUMNS + [EXCHANGE_COL]
DROP_COLUMNS = ["中文貨名", "英文貨名"]
ROC_DATE_PATTERN = re.compile(r"^\s*(\d{2,3})年\s*(\d{1,2})月\s*(?:(\d{1,2})日)?\s*$")
CANDIDATE_ENCODINGS = ["utf-8-sig", "utf-8", "big5", "cp950", "gbk"]

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)

# Per-column Excel number formats / kinds, used by build_excel.
COLUMN_FORMAT_RULES = {
    "日期": {"kind": "date", "format": "yyyy/mm/dd"},
    "貨品號列": {"kind": "int", "format": "0"},
    "重量(公斤)": {"kind": "float", "format": "#,##0"},
    "新臺幣(千元)": {"kind": "float", "format": "#,##0"},
    EXCHANGE_COL: {"kind": "float", "format": "0.0000"},
}
COLUMN_WIDTHS = {
    "進出口別": 10,
    "日期": 12,
    "貨品號列": 14,
    "重量(公斤)": 16,
    "新臺幣(千元)": 16,
    EXCHANGE_COL: 12,
}


# --------------------------------------------------------------------------
# Reading
# --------------------------------------------------------------------------

def read_csv_robust(uploaded_file) -> pd.DataFrame:
    """Try a sequence of common Taiwanese-government CSV encodings."""
    raw = uploaded_file.getvalue()
    last_err = None
    for enc in CANDIDATE_ENCODINGS:
        try:
            return pd.read_csv(io.BytesIO(raw), encoding=enc, dtype=str)
        except (UnicodeDecodeError, UnicodeError) as e:
            last_err = e
            continue
    # Last resort: decode with errors replaced so the app never hard-fails.
    return pd.read_csv(io.BytesIO(raw), encoding="utf-8", dtype=str, encoding_errors="replace")


# --------------------------------------------------------------------------
# Transform helpers (all vectorized for speed)
# --------------------------------------------------------------------------

def convert_roc_dates(series: pd.Series) -> pd.Series:
    """Vectorized conversion of 'YYY年MM月[DD日]' (ROC calendar) -> real datetime64."""
    extracted = series.astype(str).str.extract(ROC_DATE_PATTERN)
    roc_year = pd.to_numeric(extracted[0], errors="coerce")
    month = pd.to_numeric(extracted[1], errors="coerce")
    day = pd.to_numeric(extracted[2], errors="coerce").fillna(1)

    gregorian_year = roc_year + 1911
    parts = pd.DataFrame({"year": gregorian_year, "month": month, "day": day})
    valid = parts.notna().all(axis=1)

    result = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")
    if valid.any():
        result.loc[valid] = pd.to_datetime(parts.loc[valid].astype(int), errors="coerce")

    # Fallback: if a value didn't match the ROC pattern, try normal date parsing
    # so already-converted / differently formatted dates aren't lost.
    unresolved = result.isna() & series.notna()
    if unresolved.any():
        result.loc[unresolved] = pd.to_datetime(series[unresolved], errors="coerce")

    return result


def to_clean_int(series: pd.Series) -> pd.Series:
    """Strip everything non-digit (dots, dashes, spaces) and cast to nullable Int64."""
    digits_only = series.astype(str).str.replace(r"\D", "", regex=True)
    return pd.to_numeric(digits_only, errors="coerce").astype("Int64")


def to_clean_number(series: pd.Series) -> pd.Series:
    """Strip thousands separators and cast to a proper float/int."""
    cleaned = series.astype(str).str.replace(",", "", regex=False).str.strip()
    return pd.to_numeric(cleaned, errors="coerce")


def process_single_file(df: pd.DataFrame) -> pd.DataFrame:
    # 1. Drop the Chinese/English product-name columns if present.
    df = df.drop(columns=[c for c in DROP_COLUMNS if c in df.columns], errors="ignore")

    # Sanity check: make sure the columns we need actually exist.
    missing = [c for c in FINAL_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Missing expected column(s): {', '.join(missing)}")

    # 2. Reorder columns (this also drops any stray extra columns).
    df = df[FINAL_COLUMNS].copy()

    # 3. Replace 進口/出口 with IMPORT/EXPORT (substring-safe, characters don't overlap).
    df["進出口別"] = (
        df["進出口別"].astype(str).str.replace("進口", "IMPORT", regex=False)
        .str.replace("出口", "EXPORT", regex=False)
    )

    # 4. Convert ROC dates to real datetimes.
    df["日期"] = convert_roc_dates(df["日期"])

    # 5. Force 貨品號列 to a genuine number.
    df["貨品號列"] = to_clean_int(df["貨品號列"])

    # 6. Clean numeric columns.
    df["重量(公斤)"] = to_clean_number(df["重量(公斤)"])
    df["新臺幣(千元)"] = to_clean_number(df["新臺幣(千元)"])

    return df


def process_all_files(uploaded_files) -> pd.DataFrame:
    frames = [process_single_file(read_csv_robust(f)) for f in uploaded_files]
    combined = pd.concat(frames, ignore_index=True)

    # Sort by year-month only (day/time is ignored for ordering purposes).
    # Within the same year-month, the only reordering applied is pulling
    # IMPORT rows above EXPORT rows; everything else keeps its original
    # relative order (stable sort), including when only one file is uploaded.
    year_month_key = combined["日期"].dt.to_period("M")
    import_export_key = combined["進出口別"].map({"IMPORT": 0, "EXPORT": 1}).fillna(2)

    combined = (
        combined.assign(_ym=year_month_key, _ie=import_export_key)
        .sort_values(by=["_ym", "_ie"], ascending=[True, True], kind="mergesort")
        .drop(columns=["_ym", "_ie"])
        .reset_index(drop=True)
    )

    return combined


def get_months_present(df: pd.DataFrame) -> list:
    """Return the distinct year-month periods present in the 日期 column, sorted ascending."""
    return sorted(df["日期"].dt.to_period("M").dropna().unique())


def apply_exchange_rates(df: pd.DataFrame, rate_by_period: dict) -> pd.DataFrame:
    """Attach the 匯率 column: every row gets the rate entered for its row's year-month."""
    result = df.copy()
    period_key = result["日期"].dt.to_period("M")
    result[EXCHANGE_COL] = period_key.map(rate_by_period).astype(float)
    return result[OUTPUT_COLUMNS]


# --------------------------------------------------------------------------
# Clipboard export (tab-separated, no header row)
# --------------------------------------------------------------------------

def build_clipboard_tsv(df: pd.DataFrame) -> str:
    """
    Build a tab-separated block matching the final column order (A-G, including
    the exchange rate), excluding the header row, so it can be pasted directly
    into Excel starting at any cell and have Excel auto-recognize numbers/dates.
    """
    export_df = df.copy()
    export_df["日期"] = export_df["日期"].dt.strftime("%Y/%m/%d")

    # Render numbers as plain digits (no thousands separators) so Excel
    # parses them as real numeric values on paste, not text.
    numeric_cols = ["貨品號列", "重量(公斤)", "新臺幣(千元)", EXCHANGE_COL]
    for col in numeric_cols:
        export_df[col] = export_df[col].apply(
            lambda v: "" if pd.isna(v) else str(int(v)) if float(v).is_integer() else str(v)
        )

    export_df = export_df.fillna("")
    lines = ["\t".join(map(str, row)) for row in export_df.itertuples(index=False)]
    return "\n".join(lines)


def render_copy_button(tsv_text: str, row_count: int) -> None:
    """Render a button that copies `tsv_text` to the browser clipboard via JS."""
    payload = json.dumps(tsv_text)  # safely escapes quotes/newlines/backslashes for JS
    html = f"""
    <div style="display:flex; align-items:center; gap:10px; font-family:Arial, sans-serif;">
      <button id="copy-btn" style="
          background-color:#1F4E78; color:white; border:none; border-radius:6px;
          padding:0.5em 1em; font-size:14px; cursor:pointer;">
        📋 Copy {row_count:,} rows to clipboard
      </button>
      <span id="copy-status" style="font-size:13px; color:#2e7d32;"></span>
    </div>
    <script>
      const data = {payload};
      const btn = document.getElementById("copy-btn");
      const status = document.getElementById("copy-status");
      btn.addEventListener("click", async () => {{
        try {{
          await navigator.clipboard.writeText(data);
          status.textContent = "Copied! Paste into Excel with Ctrl+V / Cmd+V.";
        }} catch (err) {{
          status.textContent = "Copy failed — your browser may block clipboard access here.";
        }}
        setTimeout(() => {{ status.textContent = ""; }}, 4000);
      }});
    </script>
    """
    components.html(html, height=50)


# --------------------------------------------------------------------------
# Excel export
# --------------------------------------------------------------------------

def build_excel(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Data"

    columns = list(df.columns)

    # Header row
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, (col_name, value) in enumerate(zip(columns, row), start=1):
            rule = COLUMN_FORMAT_RULES.get(col_name)
            if rule is None:
                cell_value = None if pd.isna(value) else value
            elif pd.isna(value):
                cell_value = None
            elif rule["kind"] == "date":
                cell_value = value.to_pydatetime()
            elif rule["kind"] == "int":
                cell_value = int(value)
            else:  # float
                cell_value = float(value)

            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.font = BODY_FONT
            if rule is not None:
                cell.number_format = rule["format"]

    # Freeze header row
    ws.freeze_panes = "A2"

    # Autofit-ish column widths
    for col_idx, col_name in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        width = COLUMN_WIDTHS.get(col_name)
        if width is None:
            width = max(10, min(18, int(df[col_name].astype(str).str.len().max() or 10) + 2))
        ws.column_dimensions[letter].width = width

    ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# --------------------------------------------------------------------------
# Streamlit UI
# --------------------------------------------------------------------------

st.set_page_config(page_title="Taiwan Trade Data Merger", layout="centered")
st.title("📦 Taiwan Trade CSV → Formatted Excel")
st.write(
    "Upload one or more customs trade CSV files. The app will clean, reorder, "
    "translate, convert dates, and merge them into a single ready-to-use Excel file "
    "with an exchange-rate column you fill in per month."
)

uploaded_files = st.file_uploader(
    "Upload CSV file(s)", type=["csv"], accept_multiple_files=True
)

if uploaded_files:
    if st.button("Process files", type="primary"):
        try:
            with st.spinner("Processing..."):
                st.session_state["processed_df"] = process_all_files(uploaded_files)
            # A fresh set of processed data invalidates any previously generated output.
            st.session_state.pop("final_excel", None)
            st.session_state.pop("final_tsv", None)
        except Exception as e:
            st.error(f"Something went wrong: {e}")

    processed_df = st.session_state.get("processed_df")

    if processed_df is not None:
        st.success(f"Processed {len(processed_df):,} rows from {len(uploaded_files)} file(s).")
        st.dataframe(processed_df.head(50), use_container_width=True)

        months = get_months_present(processed_df)
        st.subheader("💱 Exchange rate per month")
        st.write("Enter the exchange rate to apply to every row in each month below.")

        rate_by_period = {}
        cols = st.columns(min(4, len(months))) if months else []
        for i, period in enumerate(months):
            with cols[i % len(cols)]:
                rate_by_period[period] = st.number_input(
                    f"{period.strftime('%Y/%m')}",
                    min_value=0.0,
                    value=0.0,
                    step=0.0001,
                    format="%.4f",
                    key=f"rate_{period}",
                )

        rates_missing = any(r <= 0 for r in rate_by_period.values())
        if rates_missing:
            st.warning("Enter a non-zero exchange rate for every month listed above.")

        if st.button("Generate Excel & clipboard data", disabled=rates_missing):
            with st.spinner("Building output..."):
                final_df = apply_exchange_rates(processed_df, rate_by_period)
                st.session_state["final_excel"] = build_excel(final_df)
                st.session_state["final_tsv"] = build_clipboard_tsv(final_df)
                st.session_state["final_row_count"] = len(final_df)

        if "final_excel" in st.session_state:
            st.download_button(
                label="⬇️ Download formatted Excel file",
                data=st.session_state["final_excel"],
                file_name=f"trade_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            st.markdown("**Or copy the data directly (no header row):**")
            render_copy_button(st.session_state["final_tsv"], st.session_state["final_row_count"])
else:
    st.info("Waiting for CSV file(s) to be uploaded.")
